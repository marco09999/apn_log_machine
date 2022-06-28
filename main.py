import Header
from init import *
from openpyxl import Workbook
from datetime import datetime
import os
from tkinter import messagebox




for machines in apn_mach:
    step_3 = []
    if machines["status"] and machines["type"] == "GROB":
        # Open text file and put each line in a list
        with open(machines["path"], "r") as tf:
            line_list = tf.read().split("\n")
            tf.close()
        # Step 1: Transform START and END BLOC into lines
        step_1 = []
        head_1 = ["Status", "Date", "Time", "Pallet", "Name", "Post Version", "Post Date", "Force Version", "Force Date", "Otr",
                  "Line in text file"]
        step_1.append(head_1)
        for i in range(0, len(line_list)):
            if len(Header.line_decode_step_1(line_list, i)) > 1:
                step_1.append(Header.line_decode_step_1(line_list, i))
        # Step 2: Transform START and END lines into execution
        step_2 = []
        head_2 = ["Name", "Cycle time (min)", "Post", "Force", "Otr", "Start Date", "Start Time", "End Date", "End time"]
        step_2.append(head_2)
        for i in range(0, len(step_1)-1):
            if len(Header.line_decode_step_2(step_1, i)) > 1:
                step_2.append(Header.line_decode_step_2(step_1, i))
        # Step 2.5: Add identifier at the end of each execution
        step_2_5 = step_2
        head_2_5 = ["Name", "Cycle time (min)", "Post", "Force", "Otr", "Start Date", "Start Time", "End Date",
                    "End time", "Identifier"]
        step_2_5[0] = head_2_5
        for i in range(1, len(step_2_5)):
            step_2[i].append(step_2[i][0] + "-" + step_2[i][2] + "-" + step_2[i][3] + "-" + step_2[i][4])
        # Step 3: Compress into single line and number of execution for each unique program
        head_3 = ["Program Name", "Cycle time (min)", "Execution", "Start Date", "Start time", "End date", "End time",
                  "Post", "Force", "Otr", "Machine"]
        step_3 = Header.compress(step_2_5)
        # Add machine Number
        for i in range(0, len(step_3)):
            if i != 0:
                step_3[i].append(machines["name"])
    machines["log"] = step_3

# Create an Excel Workbook
wb = Workbook()

# For each machine, create a sheet and write step_3 to the sheet
for mach_dico in apn_mach:
    ws = wb.create_sheet(mach_dico["name"])
    full_log = mach_dico["log"]
    for i in range(0, len(full_log)):
        time = "LAST UPDATE: " + datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws.cell(row=1, column=12, value=time)
        actual_line = full_log[i]
        row = i+1
        for j in range(0, len(actual_line)):
            actual_value = actual_line[j]
            col = j+1
            ws.cell(row=row, column=col, value=actual_value)

try:
    wb.save("C:\Temp\LOG_MACHINE.xlsx")
    os.startfile("C:\Temp\LOG_MACHINE.xlsx")
except:
    print("test")







