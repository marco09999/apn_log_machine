import statistics
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def get_date(line):
    if "TIME-" in line:
        start = line.index("TIME-") + 5
        end = line.index(" - ")
        return line[start:end]
    else:
        return "UNABLE TO FIND DATE"


def get_time(line):
    if "TIME-" in line:
        start = line.index(" - ") + 3
        end = len(line) + 1
        return line[start:end]
    else:
        return "UNABLE TO FIND TIME"


def get_pallet(line):
    if line == str(1) or line == str(2):
        return line
    else:
        return "UNABLE TO FIND PALLET"


def get_name(line):
    if ".spf" in line or ".SPF" in line:
        return line
    else:
        return "UNABLE TO FIND NAME"


def get_post_version(line):
    if "OPTION FILE" in line:
        start = line.index("OPTION FILE") + 14
        end = len(line)
        return line[start:end]
    else:
        return "UNABLE TO FIND POST VERSION"


def get_post_date(line):
    if "DATE :" in line:
        start = line.index("DATE :") + 7
        end = len(line)
        return line[start:end]
    else:
        return "UNABLE TO FIND POST DATE"


def get_force_version(line):
    start = line.index("VERSION") + 8
    end = start + 5
    return line[start:end]


def get_force_date(line):
    day = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    for i in day:
        if i in line:
            start = line.index(i)
            end = line.index("202") + 4
    return line[start:end]


def get_otr(line):
    start = line.index("VERSION") + 8
    end = len(line)
    return line[start:end]


def line_decode_step_1(list, index):
    # Format of step_1 is : Status, Date, Time, Pallet, Name, Post Version, Post Date, Force Version, Force Date, Otr
    step_1 = []
    shift = 0
    is_force_opti = False
    is_retract_opti = False

    if "START" in list[index]:
        if list[index].index("START") == 0:
            step_1.append("START")                                              # Status
            step_1.append(get_date(list[index + 1]))                            # Date
            step_1.append(get_time(list[index + 1]))                            # Time
            step_1.append(get_pallet(list[index + 2]))                          # Pallet

            for i in range(0, 16):
                line = list[index + i + 4]
                if "VERICUT" in line:
                    force_version = get_force_version(line)
                    force_date = get_force_date(list[index + i + 4 + 2])
                    is_force_opti = True
                    shift += 7
                elif "OPTI_TOOL" in line:
                    otr_version_date = get_otr(line)
                    is_retract_opti = True
                    shift += 1

            step_1.append(get_name(list[index + 3]))                    # Name
            step_1.append(get_post_version(list[index + 9 + shift]))    # Post Version
            step_1.append(get_post_date(list[index + 8 + shift]))       # Post Date

            if not(is_force_opti):
                step_1.append("-")                      #Force Version if not opti
                step_1.append("-")                      #Force Date if not opti
            if not(is_retract_opti):
                step_1.append("-")                      #Otr if not opti
            if is_force_opti:
                step_1.append(force_version)            #Force Version if opti
                step_1.append(force_date)               #Force Version if opti
            if is_retract_opti:
                step_1.append(otr_version_date)         #otr if opti
            step_1.append(index + 1)

    if "*/*--STOPSTART" in list[index]:
        step_1.append("STOPSTART")

    if "END" in list[index]:
        if list[index].index("END") == 0:
            step_1.append("END")
            step_1.append(get_date(list[index + 1]))
            step_1.append(get_time(list[index + 1]))
            step_1.append(get_pallet(list[index+ 2]))
            step_1.append(get_name(list[index + 3]))
            step_1.append(index + 1)
    if "*/*--STOPEND" in list[index]:
        step_1.append("STOPEND")
    return step_1


def cycle_time_step_2(start_list,end_list):
    start_time_str = start_list[2]
    end_time_str = end_list[2]
    start_time_h = int(start_time_str[0:start_time_str.index(":")])
    start_time_m = int(start_time_str[start_time_str.index(":")+1:start_time_str.rindex(":")])
    start_time_s = int(start_time_str[start_time_str.rindex(":")+1: len(start_time_str)])
    end_time_h = int(end_time_str[0:end_time_str.index(":")])
    end_time_m = int(end_time_str[end_time_str.index(":")+1: end_time_str.rindex(":")])
    end_time_s = int(end_time_str[end_time_str.rindex(":")+1: len(end_time_str)])
    start_date = start_list[1]
    end_date = end_list[1]
    if start_date != end_date:
        hr = end_time_h - start_time_h + 24
        min = end_time_m - start_time_m
        sec = end_time_s - start_time_s
    else:
        hr = end_time_h - start_time_h
        min = end_time_m - start_time_m
        sec = end_time_s - start_time_s

    return round((hr*60) + min + (sec/60), 2)


def line_decode_step_2(list, index):
    step_2 = []
    status_start = list[index][0]
    status_end = list[index + 1][0]
    pal_start = list[index][3]
    pal_stop = list[index + 1][3]
    name_start = list[index][4]
    name_stop = list[index + 1][4]
    if status_start == "START" and status_end == "END" and pal_start == pal_stop and name_start == name_stop:
        step_2.append(name_start)
        step_2.append(cycle_time_step_2(list[index], list[index + 1]))
        step_2.append(list[index][5] + "-" + list[index][6])
        step_2.append(list[index][7] + "-" + list[index][8])
        step_2.append(list[index][9])
        step_2.append(list[index][1])
        step_2.append(list[index][2])
        step_2.append(list[index + 1][1])
        step_2.append(list[index + 1][2])
    return step_2


def compress(list):
    execution_list = []
    head = ["Program Name", "Cycle time (min)", "Execution", "Start Date", "Start time", "End date", "End time",
            "Post", "Force", "Otr", "Machine"]
    execution_list.append(head)
    for i in range(1, len(list)):
        execution = []
        name = list[i][0]
        post = list[i][2]
        force = list[i][3]
        otr = list[i][4]
        cycle_time = []
        start_date = []
        start_time = []
        end_date = []
        end_time = []
        identifier1 = list[i][9]
        if identifier1 != "read":
            cycle_time.append(list[i][1])
            start_date.append(list[i][5])
            start_time.append(list[i][6])
            end_date.append(list[i][7])
            end_time.append(list[i][8])
            list[i][9] = "read"
            for j in range(i, len(list)):
                identifier2 = list[j][9]
                if identifier2 == identifier1 and identifier1 != "read":
                    cycle_time.append(list[j][1])
                    start_date.append(list[j][5])
                    start_time.append(list[j][6])
                    end_date.append(list[j][7])
                    end_time.append(list[j][8])
                    list[j][9] = "read"
            nb_exec = len(cycle_time)
            if nb_exec > 8:
                ans = round(statistics.median(cycle_time), 2)
                execution.append(name)
                execution.append(ans)
                execution.append(nb_exec)
                execution.append(start_date[0])
                execution.append(start_time[0])
                execution.append(end_date[len(end_date) - 1])
                execution.append(end_time[len(end_time) - 1])
                execution.append(post)
                execution.append(force)
                execution.append(otr)
                execution_list.append(execution)
            if nb_exec <= 8:
                ans = str(round(statistics.mean(cycle_time), 2)) + " *!NA!*"
                execution.append(name)
                execution.append(ans)
                execution.append(nb_exec)
                execution.append(start_date[0])
                execution.append(start_time[0])
                execution.append(end_date[len(end_date) - 1])
                execution.append(end_time[len(end_time) - 1])
                execution.append(post)
                execution.append(force)
                execution.append(otr)
                execution_list.append(execution)

    return execution_list


def format_data(ws):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 17.4
    ws.column_dimensions['C'].width = 10.43
    ws.column_dimensions['D'].width = 9.43
    ws.column_dimensions['E'].width = 9.57
    ws.column_dimensions['F'].width = 9.43
    ws.column_dimensions['G'].width = 9.57
    ws.column_dimensions['H'].width = 38.14
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 11
    ws.column_dimensions['L'].width = 38
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=4).font = Font(bold=True)
    ws.cell(row=1, column=5).font = Font(bold=True)
    ws.cell(row=1, column=6).font = Font(bold=True)
    ws.cell(row=1, column=7).font = Font(bold=True)
    ws.cell(row=1, column=8).font = Font(bold=True)
    ws.cell(row=1, column=9).font = Font(bold=True)
    ws.cell(row=1, column=10).font = Font(bold=True)
    ws.cell(row=1, column=11).font = Font(bold=True)
    ws.cell(row=1, column=12).font = Font(bold=True)
    return






